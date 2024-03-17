import openpyxl
from openpyxl.styles import Font, Color, PatternFill
from copy import copy

thoi_gian = ['0600 - 0615', '0615 - 0630', '0630 - 0645', '0645 - 0700',
             '0700 - 0715', '0715 - 0730', '0730 - 0745', '0745 - 0800',
             '0800 - 0815', '0815 - 0830', '0830 - 0845']

#m c b l h
so_luong_xe = [[178, 165, 56, 89, 56], [145, 185, 56, 89, 56], [162, 133, 56, 89, 56], [176, 171, 56, 89, 56],
                 [100, 120, 56, 89, 56], [90, 110, 56, 89, 56], [80, 100, 56, 89, 56], [70, 90, 56, 89, 56],
                 [200, 220, 56, 89, 56], [210, 230, 56, 89, 56], [195, 205, 56, 89, 56]]
#
def xxuat_xlsx(path_xlsx, thoi_gian,so_luong_xe, name_proj):
    so_luong_xe1 = [sublist[:5] for sublist in so_luong_xe]
    print(so_luong_xe[:5][:])
    for i in range(len(so_luong_xe)):
        so_luong_xe1[i][0] = so_luong_xe[i][1]
        so_luong_xe1[i][1] = so_luong_xe[i][3]
        so_luong_xe1[i][2] = so_luong_xe[i][4]
        so_luong_xe1[i][3] = so_luong_xe[i][0]
        so_luong_xe1[i][4] = so_luong_xe[i][2]
    so_luong_xe = so_luong_xe1

    # Đường dẫn đến file kết quả
    file_ket_qua = 'data_mau.xlsx'

    wb = openpyxl.load_workbook(file_ket_qua)

    # Chọn sheet cần thao tác
    sheet = wb.active
    empty_rows=[]
    # Duyệt qua các dòng trong cột A
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        for i, cell_value in enumerate(row, start=1):
            source_cell = sheet.cell(row=row_index, column=i)
            # print(type(source_cell))
            # print(source_cell)
            source_style = copy(source_cell._style)
            # print(source_style)
            if cell_value in thoi_gian:
                index = thoi_gian.index(cell_value)
                so_luong = so_luong_xe[index]
                # Điền số lượng vào các ô tiếp theo cùng dòng
                for j, sl in enumerate(so_luong):
                    cell = sheet.cell(row=row_index, column=i+1+j)
                    cell.value = sl
                    cell._style = source_style

    cell_A9 = sheet.cell(row=9, column=1)
    cell_A9.value = name_proj
    cell_B10 = sheet.cell(row=10, column=2)
    cell_B10.value = "V1"
    # Lưu file kết quả
    wb.save(path_xlsx)
    print("lưu:", path_xlsx)

def xxuat_xlsx_9class(path_xlsx, thoi_gian,so_luong_xe, name_proj):
    so_luong_xe1 = [sublist[:9] for sublist in so_luong_xe]
    print(so_luong_xe[:9][:])
    # for i in range(len(so_luong_xe)):
    #     so_luong_xe1[i][0] = so_luong_xe[i][1]
    #     so_luong_xe1[i][1] = so_luong_xe[i][3]
    #     so_luong_xe1[i][2] = so_luong_xe[i][4]
    #     so_luong_xe1[i][3] = so_luong_xe[i][0]
    #     so_luong_xe1[i][4] = so_luong_xe[i][2]
    so_luong_xe = so_luong_xe1

    # Đường dẫn đến file kết quả
    file_ket_qua = 'result_9class.xlsx'

    wb = openpyxl.load_workbook(file_ket_qua)

    # Chọn sheet cần thao tác
    sheet = wb.active
    empty_rows=[]
    # Duyệt qua các dòng trong cột A
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        for i, cell_value in enumerate(row, start=1):
            source_cell = sheet.cell(row=row_index, column=i)
            # print(type(source_cell))
            # print(source_cell)
            source_style = copy(source_cell._style)
            # print(source_style)
            if cell_value in thoi_gian:
                index = thoi_gian.index(cell_value)
                so_luong = so_luong_xe[index]
                # Điền số lượng vào các ô tiếp theo cùng dòng
                for j, sl in enumerate(so_luong):
                    cell = sheet.cell(row=row_index, column=i+1+j)
                    cell.value = sl
                    cell._style = source_style

    cell_A9 = sheet.cell(row=9, column=1)
    cell_A9.value = name_proj
    cell_B10 = sheet.cell(row=10, column=2)
    cell_B10.value = "V1"
    # Lưu file kết quả
    wb.save(path_xlsx)
    print("lưu:", path_xlsx)

